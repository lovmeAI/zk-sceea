from datetime import datetime, date, time

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet


class Excel:
    def _get_model_fields(self, model, primary_key=False, comment=False):
        """
        根据SQLAlchemy模型生成需要的数据
        :param model: SQLAlchemy模型
        :param comment: 导出数据的时候用
        :return:
        """
        title = ''
        fields = {}
        index = 0
        for key, value in model.__dict__.items():
            if key == '__table_args__':
                title = value.get('comment')
                fields[title] = {}
            if key.startswith('_'): continue
            if primary_key and value.primary_key: continue
            if comment:
                fields[title][key] = {'field': value.comment, 'comment': None}
            else:
                fields[title][self._column(index)] = {'field': key, 'type': self._sql_type(value.type.__class__.__name__)}
                index += 1
        return fields

    def _sql_type(self, _type):
        temp = {"String": str, "Boolean": bool, "DateTime": datetime, "Integer": int, "Float": float, "DECIMAL": float, "Date": date, "Time": time, "Text": str, "LONGTEXT": str}
        return temp.get(_type, str)

    def _column(self, num: int):
        """
        生成Excel字段
        :param num: 0-25 A-Z 最大 0-701 A-ZZ
        :return: str
        """
        if num > 701:
            raise Exception("0-25 A-Z 最大最小区间 0-701 A-ZZ， 最多生成701个字段，比这还多，自己增加去吧")
        multiple, remainder = divmod(num, 26)
        if multiple > 0:
            letter = '%s%s' % (chr(multiple + 64), chr(remainder + 65))
        else:
            letter = chr(num + 65)
        return letter


class ExcelRead(Excel):
    """
    fields = {"Sheet名称": ['group_name', 'enable', 'remark'], 'default': ['group_name', 'content', 'enable', 'level']}
    excel = ExcelRead('./策略信息-20210519103521.xlsx', field=fields)
    excel = ExcelRead('./策略信息-20210519103521.xlsx', models=Model)
    excel = ExcelRead('./策略信息-20210519103521.xlsx', models=[Model, Model])
    for i in excel.data:
        print(i)
    """

    def __init__(self, file, field: dict = None, is_title=True, start: int = 1, sheet_name: [list, str] = None):
        """
        Excel 文件读取
        models 和 field只能使用一个，models > field
        :param file: Excel文件路径或Excel对象
        :param field: 配置数据字段 {"策略组": ['group_name', 'enable', 'remark'], 'default': ['group_name', 'content', 'enable', 'level']}
        :param start: 从第 start 行开始读取数据
        :param sheet_name: 读取指定的sheet，一个或多个
        """
        self._start = start
        self._sheet_names = sheet_name
        self._field = {}
        self._param_field = field
        self._is_title = is_title
        self._excel = load_workbook(file)
        if field:
            for key, val in field.items():
                temp = {}
                for i, value in enumerate(val):
                    temp[self._column(i)] = value
                self._field[key] = temp

    def _read_sheet(self, sheet: Worksheet, sheet_name, start) -> list:
        """解析sheet工作表数据"""
        field_dict = self._field.get(sheet_name)
        if not field_dict and self._param_field:
            field_dict = self._field.get('default')
        field_dict = field_dict if field_dict else {}
        sheet_rows = list(sheet.rows)
        if len(self._field.keys()) <= 0 and self._is_title:
            for cell in sheet_rows[0]:
                letter = cell.column_letter
                field_dict[letter] = cell.value
        temp_list = []
        for _, row in enumerate(sheet_rows[start:]):
            temp = {}
            for index, cell in enumerate(row):
                letter = cell.column_letter
                field = field_dict.get(letter, letter)
                if isinstance(field, dict):
                    temp[field['field']] = cell.value
                else:
                    temp[field] = cell.value
            temp_list.append(temp)
        yield sheet_name, list(field_dict.values()), temp_list

    @property
    def data(self) -> Worksheet:
        """解析excel文件并获取sheet"""
        if self._sheet_names:
            if isinstance(self._sheet_names, list):
                for sheet in self._sheet_names:
                    return next(self._read_sheet(self._excel[sheet], sheet, self._start))
            elif isinstance(self._sheet_names, str):
                return next(self._read_sheet(self._excel[self._sheet_names], self._sheet_names, self._start))
            else:
                raise Exception("传入的 excel_read 方法的 sheet_name 参数 不是一个list或者str")
        else:
            for sheet in self._excel.sheetnames:
                return next(self._read_sheet(self._excel[sheet], sheet, self._start))

    def __del__(self):
        self._excel.close()


class ExcelWrite(Excel):
    """
    header = {"策略组": {
        '策略组名称': '必填\n说明:\n策略组名称长度应不超过64个字符\n策略组名称应互不相同\n若新添加的策略组名称和已有策略组冲突，则前者将覆盖后者',
        '状态':    '必填\n说明:\n值可以是数字0或者1\n1表示启用策略组\n0表示禁用策略组',
        '备注':    '选填\n说明:\n备注信息长度应当不超过256个字符'
    }, 'default':    {
        '策略组名称': '必填\n说明:\n请填写正确的策略组名称\n策略组名称长度应不超过64个字符',
        '策略组内容': '必填\n说明:\n策略内容长度应不超过256个字符',
        '状态':    '必填\n说明:\n值可以是数字0或者1\n1表示启用策略\n0表示禁用策略',
        '保密级别':  '必填\n说明:\n敏感等级的值可以是数字1,2,3,4,5，数字大小从小到大表示敏感等级逐渐升高，依次对应非密、内部、秘密、机密、绝密'
    }}
    header2 = {"策略组": {
        'field1': '必填\n说明:\n策略组名称长度应不超过64个字符\n策略组名称应互不相同\n若新添加的策略组名称和已有策略组冲突，则前者将覆盖后者',
        'field2':    '必填\n说明:\n值可以是数字0或者1\n1表示启用策略组\n0表示禁用策略组',
        'field3':    '选填\n说明:\n备注信息长度应当不超过256个字符'
    }, 'default':    {
        '策略组名称': '必填\n说明:\n请填写正确的策略组名称\n策略组名称长度应不超过64个字符',
        '策略组内容': '必填\n说明:\n策略内容长度应不超过256个字符',
        '状态':    '必填\n说明:\n值可以是数字0或者1\n1表示启用策略\n0表示禁用策略',
        '保密级别':  '必填\n说明:\n敏感等级的值可以是数字1,2,3,4,5，数字大小从小到大表示敏感等级逐渐升高，依次对应非密、内部、秘密、机密、绝密'
    }}
    write = ExcelWrite("./xxxx.xlsx", header) # 实例化
    write = ExcelWrite("./xxxx.xlsx", header) # 实例化
    write.create_sheet("Sheet名称", [{},{},{},...])
    write.save()
    """

    def __init__(self, filename: str, models: [list, str] = None, hidden_col: dict = None, hidden_sheet: [list, int, str] = None, header: dict = None, author='admin'):
        """
        创建Excel工作簿
        如果 models 存在，header={"Sheet名称": {'field1': '备注','field2': '备注','field3': '备注'}, field 为数据库字段名称
        如果 models 不存在，header={"Sheet名称": {'自定义1': '备注','自定义2': '备注','自定义3': '备注'}, field 为数据库字段名称
        :param filename: 文件路径
        :param models: SQLAlchemy 模型对象
        :param hidden_sheet: 隐藏指定的sheet number(数字)|sheet名称 or [下标1,下标2,下标3,...]|['Sheet名称','Sheet名称2',...]  number(数字)是下标 从0开始
        :param hidden_col: 隐藏指定的列 {"Sheet名称": ['A','B','C',...], "Sheet名称2": ['A','B',...]} or {"Sheet名称": 'A', "Sheet名称2": 'B'}
        :param header: 表头{"Sheet名称": {'字段1': '备注','字段2': '备注','字段3': '备注'}, 'default': {'字段1': '备注','字段2': '备注','字段3': '备注'}}
                       {"Sheet名称": ['字段1','字段2','字段3'], 'default': ['字段1','字段2','字段3']} 字段 为自定义字段，Excel的表头
        :param author: 作者
        """
        self._filename = filename
        self._excel = Workbook()
        self._header = {}
        self._author = author
        self._hidden_col = hidden_col
        self._hidden_sheet = hidden_sheet
        self._sheets = []
        if models:
            if isinstance(models, list):
                for model in models:
                    self._header.update(self._get_model_fields(model, comment=True))
            else:
                self._header.update(self._get_model_fields(models, comment=True))
            _header = {}
            for key, val in self._header.items():
                _header[key] = {}
                for k, v in val.items():
                    if header and header.get('default', None):
                        if header['default'].get(k, None):
                            self._header[key][k]['comment'] = header['default'][k]
                    if header and header.get(key, None):
                        if header[key].get(k, None):
                            self._header[key][k]['comment'] = header[key][k]
                    _header[key][v['field']] = v['comment']
            self._header = _header
        if not models and header:
            self._header = header

    def create_sheet(self, sheet_name: str, data: list):
        """
        创建Sheet和Sheet中的数据
        :param sheet_name: Sheet名称
        :param data: Sheet里面的数据 [{},{},{},...]
        :return:
        """
        self._sheets.append(sheet_name)
        self.sheet: Worksheet = self._excel.create_sheet(sheet_name)
        header = self._header.get(sheet_name)
        header = header if header else self._header.get('default')
        # 写入表头
        if header:
            if isinstance(header, dict):
                self.sheet.append(list(header.keys()))
                for i, val in enumerate(list(header.values())):
                    if val:
                        self.sheet['%s1' % self._column(i)].comment = Comment(val, self._author)
            if isinstance(header, list):
                self.sheet.append(header)

        # 写入数据
        for item in data:
            if isinstance(item, dict):
                self.sheet.append(list(item.values()))
                continue
            elif isinstance(item, str):
                self.sheet.append(data)
                break
            else:
                self.sheet.append(list(item))

        # 隐藏指定列
        if self._hidden_col:
            hidden = self._hidden_col.get(sheet_name)
            hidden = hidden if hidden else self._hidden_col.get('default')
            if isinstance(hidden, str):
                self.sheet.column_dimensions[hidden.upper()].hidden = True
            if isinstance(hidden, list):
                for i in hidden:
                    self.sheet.column_dimensions[i.upper()].hidden = True

    def _is_hidden_sheets(self):
        """隐藏指定的sheet"""
        if self._hidden_sheet:
            if isinstance(self._hidden_sheet, str) and self._hidden_sheet in self._sheets:
                self._excel[self._hidden_sheet].sheet_state = 'hidden'
            elif isinstance(self._hidden_sheet, int):
                self._excel[self._sheets[self._hidden_sheet]].sheet_state = 'hidden'
            elif isinstance(self._hidden_sheet, list):
                if isinstance(self._hidden_sheet[0], str):
                    for sheet in self._hidden_sheet:
                        if sheet in self._sheets:
                            self._excel[sheet].sheet_state = 'hidden'
                elif isinstance(self._hidden_sheet[0], int):
                    for i in self._hidden_sheet:
                        self._excel[self._sheets[i]].sheet_state = 'hidden'

    def save(self):
        """保存文件"""
        self._excel.remove(self._excel["Sheet"])
        self._is_hidden_sheets()
        self._excel.save(self._filename)

    def __del__(self):
        self._excel.close()


if __name__ == '__main__':
    fields = {"策略组": ['group_name', 'enable', 'remark'], 'default': ['group_name', 'content', 'enable', 'level']}
    excel = ExcelRead("D:\\development\\project\\terminalcheck-server\\luffy\\策略信息-20210519103521.xlsx", field=fields)
    header = {"策略组": {
        '策略组名称': '必填\n说明:\n策略组名称长度应不超过64个字符\n策略组名称应互不相同\n若新添加的策略组名称和已有策略组冲突，则前者将覆盖后者',
        '状态': '必填\n说明:\n值可以是数字0或者1\n1表示启用策略组\n0表示禁用策略组',
        '备注': '选填\n说明:\n备注信息长度应当不超过256个字符'
    }, 'default': {
        '策略组名称': '必填\n说明:\n请填写正确的策略组名称\n策略组名称长度应不超过64个字符',
        '策略组内容': '必填\n说明:\n策略内容长度应不超过256个字符',
        '状态': '必填\n说明:\n值可以是数字0或者1\n1表示启用策略\n0表示禁用策略',
        '保密级别': '必填\n说明:\n敏感等级的值可以是数字1,2,3,4,5，数字大小从小到大表示敏感等级逐渐升高，依次对应非密、内部、秘密、机密、绝密'
    }}
    write = ExcelWrite("xxxx.xlsx", header)
    for i in excel.data:
        write.create_sheet(*i)
    write.save()
